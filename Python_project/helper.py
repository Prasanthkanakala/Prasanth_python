import os
import csv
from numpy import NAN,nan
from PIL.Image import linear_gradient
import pandas as pd
from openpyxl import Workbook
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font,Alignment,Border,Side, alignment

path=os.getcwd()
input=path+'/sample_input/'

os.chdir(input)

pos=5
neg=1
zero=0

master_file=pd.read_csv("master_roll.csv")
print(master_file.head())
master_roll=master_file['roll'].values.tolist()
master_name=master_file['name'].values.tolist()
#print(master_roll)


dict_roll={}
for i in range(len(master_roll)):
    dict_roll[master_roll[i]]=master_name[i]
    print(dict_roll[master_roll[i]])

response_file=pd.read_csv("responses.csv")
print(response_file.head())
response_roll=response_file['Roll Number'].values.tolist()
correct_options=response_file.iloc[0].values.tolist()
correct_options=correct_options[6:]

os.chdir(path)
if not os.path.exists('output'):
    os.mkdir('output')

status=[]
final_score=[]
no_stu_master=0
no_of_response=0
#print(response_roll)
os.chdir(path)

alignment_heading=Alignment(horizontal='right',vertical='bottom')
alignment_content=Alignment(horizontal='left',vertical='bottom')
alignment_ans=Alignment(horizontal='center',vertical='bottom')
font_heading=Font(name='Calibri',size=14,bold=False)
font_content=Font(name='Calibri',size=14,bold=True)
right_color=Font(color='00FF00',name='Calibri',size=14,bold=False)
left_color=Font(color='ff0000',name='Calibri',size=14,bold=False)
give_color=Font(color='0000FF',name='Calibri',size=14,bold=False)
border_style=Side(border_style='medium',color='000000')

while(no_stu_master<len(master_roll)):
    wb=Workbook()
    sheet=wb.active
    img=Image('iitp_logo.png')
    alignment_heading=Alignment(horizontal='right',vertical='bottom')
    alignment_content=Alignment(horizontal='left',vertical='bottom')
    alignment_ans=Alignment(horizontal='center',vertical='bottom')
    font_heading=Font(name='Calibri',size=14,bold=False)
    font_content=Font(name='Calibri',size=14,bold=True)
    right_color=Font(color='00FF00',name='Calibri',size=14,bold=False)
    left_color=Font(color='ff0000',name='Calibri',size=14,bold=False)
    give_color=Font(color='0000FF',name='Calibri',size=14,bold=False)
    border_style=Side(border_style='medium',color='000000')
    border=Border(top=border_style,bottom=border_style,left=border_style,right=border_style)
    img.width=610
    img.height=80
    sheet.add_image(img,'A1')
    sheet.column_dimensions['A'].width=17
    sheet.column_dimensions['B'].width=17
    sheet.column_dimensions['C'].width=17
    sheet.column_dimensions['D'].width=17
    sheet.column_dimensions['E'].width=17

    sheet['A6']='Name :'
    sheet['A6'].font=font_heading
    sheet['A6'].alignment=alignment_heading
    
    sheet['D6']='Exam :'
    sheet['D6'].font=font_heading
    sheet['D6'].alignment=alignment_heading
    sheet['E6']='quiz'
    sheet['E6'].font=font_content
    sheet['E6'].alignment=alignment_content
    sheet['A7']='Roll Number :'
    sheet['A7'].font=font_heading
    sheet['A7'].alignment=alignment_heading

    empty=['A9','A10','A11','A12','B9','C9','D9','E9','A15','B15','D15','E15']
    for key in empty:
        sheet[key].font=font_content
        sheet[key].alignment=alignment_ans
        sheet[key].border=border
    
    sheet.merge_cells('A5:E5')
    sheet.cell(row=5,column=1).font=Font(size=18,bold=True,name='Century',underline='single')
    sheet.cell(row=5,column=1).alignment=alignment_ans
    sheet.cell(row=5,column=1).value='Mark Sheet'

    sheet.merge_cells('B6:C6')
    sheet.cell(row=5,column=1).font=font_content
    sheet.cell(row=5,column=1).alignment=alignment_content
    sheet.cell(row=5,column=1).value=dict_roll[master_roll[no_stu_master]]

    sheet.cell(row=7,column=2).font=font_content
    sheet.cell(row=7,column=2).alignment=alignment_content
    sheet.cell(row=7,column=2).value=master_roll[no_stu_master]

    sheet.cell(row=10,column=5).font=font_content
    sheet.cell(row=10,column=5).alignment=alignment_ans
    sheet.cell(row=10,column=5).value=str(len(correct_options))

    sheet.cell(row=9,column=2).font=Font(size=12,bold=True, name='Calibri')
    sheet.cell(row=9,column=2).alignment=Alignment('center')
    sheet.cell(row=9,column=2).value='Right'

    sheet.cell(row=9,column=3).font=Font(size=12,bold=True, name='Calibri')
    sheet.cell(row=9,column=3).alignment=Alignment('center')
    sheet.cell(row=9,column=3).value='Wrong'

    sheet.cell(row=9,column=4).font=Font(size=12,bold=True, name='Calibri')
    sheet.cell(row=9,column=4).alignment=Alignment('center')
    sheet.cell(row=9,column=4).value='Not Attempt'

    sheet.cell(row=9,column=5).font=Font(size=12,bold=True, name='Calibri')
    sheet.cell(row=9,column=5).alignment=Alignment('center')
    sheet.cell(row=9,column=5).value='Max'

    sheet.cell(row=11,column=2).font=Font(size=12,color='00008000', name='Calibri')
    sheet.cell(row=11,column=2).alignment=Alignment('center')
    sheet.cell(row=11,column=2).value=str(pos)

    sheet.cell(row=11,column=2).font=Font(size=12,color='00FF0000', name='Calibri')
    sheet.cell(row=11,column=2).alignment=Alignment('center')
    sheet.cell(row=11,column=2).value=str(neg)

    sheet.cell(row=11,column=2).font=Font(size=12, name='Calibri')
    sheet.cell(row=11,column=2).alignment=Alignment('center')
    sheet.cell(row=11,column=2).value=str(zero)

    sheet.cell(row=15,column=4).font=Font(size=12,bold=True, name='Calibri')
    sheet.cell(row=15,column=4).alignment=Alignment('center')
    sheet.cell(row=15,column=4).value='Student Ans'

    sheet.cell(row=15,column=2).font=Font(size=12,bold=True, name='Calibri')
    sheet.cell(row=15,column=2).alignment=Alignment('center')
    sheet.cell(row=15,column=2).value='Correct Ans'

    sheet.cell(row=15,column=5).font=Font(size=12,bold=True, name='Calibri')
    sheet.cell(row=15,column=5).alignment=Alignment('center')
    sheet.cell(row=15,column=5).value='Correct Ans'

    line=16
    m=0
    while(m<len(correct_options)):
        if(line<=40):
            sheet.cell(row=line,column=2).font=Font(size=12,color='000000FF', name='Calibri')
            sheet.cell(row=line,column=2).alignment=Alignment('center')
            sheet.cell(row=line,column=2).value=correct_options[m]

        else:
            sheet.cell(row=line-25,column=5).font=Font(size=12,color='000000FF', name='Calibri')
            sheet.cell(row=line-25,column=5).alignment=Alignment('center')
            sheet.cell(row=line-25,column=5).value=correct_options[m]

        line+=1
        m+=1

    #set_border(sheet,'A15:B40')
    #set_border(sheet,'D15:E'+'{}'.format(line-25))

    sheet.cell(row=10,column=1).alignment=Alignment('center')
    sheet.cell(row=10,column=1).font=Font(size=12,bold=True,name="Calibri")
    sheet.cell(row=10,column=1).value='No.'

    sheet.cell(row=11,column=1).alignment=Alignment('center')
    sheet.cell(row=11,column=1).font=Font(size=12,bold=True,name="Calibri")
    sheet.cell(row=11,column=1).value='Marking'

    sheet.cell(row=12,column=1).alignment=Alignment('center')
    sheet.cell(row=12,column=1).font=Font(size=12,bold=True,name="Calibri")
    sheet.cell(row=12,column=1).value='Total'

    if(no_of_response<len(response_roll) and master_roll[no_stu_master]==response_roll[no_of_response]):

        stu_options=[]
        stu_options=response_file.iloc[no_of_response].values.tolist()
        stu_options=stu_options[6:]

        k=0
        total=0
        wrong=0
        right=0
        notattempt=0
        while(k<len(correct_options)):
            if(stu_options[k]==correct_options[k]):
                right+=1
            k+=1

        nonattempt=stu_options.count(nan)
        wrong=len(stu_options)-right-notattempt
        status.append('[{},{},{}]'.format(right,wrong,notattempt))

        total=((right*5)-(wrong*1))
        max_marks=len(correct_options)*5
        final_score.append('{}/{}'.format(total,max_marks))

        #set_border(sheet,'A9:E12')

        sheet.cell(row=10,column=2).font=Font(size=12,color="00008000",name="Calibri")
        sheet.cell(row=10,column=2).alignment=Alignment('center')
        sheet.cell(row=10,column=2).value=str(right)

        sheet.cell(row=10,column=3).font=Font(size=12,color="00FF0000",name="Calibri")
        sheet.cell(row=10,column=3).alignment=Alignment('center')
        sheet.cell(row=10,column=3).value=str(wrong)

        sheet.cell(row=10,column=4).font=Font(size=12,name="Calibri")
        sheet.cell(row=10,column=4).alignment=Alignment('center')
        sheet.cell(row=10,column=4).value=str(notattempt)

        sheet.cell(row=12,column=2).font=Font(size=12,color="00008000",name="Calibri")
        sheet.cell(row=12,column=2).alignment=Alignment('center')
        sheet.cell(row=12,column=2).value=str(right*5)

        sheet.cell(row=12,column=3).font=Font(size=12,color="00FF0000",name="Calibri")
        sheet.cell(row=12,column=3).alignment=Alignment('center')
        sheet.cell(row=12,column=3).value=str(wrong*(-1))

        sheet.cell(row=12,column=4).font=Font(size=12,name="Calibri",color='000000FF')
        sheet.cell(row=12,column=4).alignment=Alignment('center')
        sheet.cell(row=12,column=4).value='{}/{}'.format(total,max_marks)


        line=16
        m=0
        while(m<len(correct_options)):
            if(line<=40):
                if(stu_options[m]==correct_options[m]):
                    sheet.cell(row=line,column=1).font=Font(size=12,color="00008000",name="Calibri")
                    sheet.cell(row=line,column=1).alignment=Alignment('center')
                    sheet.cell(row=line,column=1).value=stu_options[m]

                else:
                    sheet.cell(row=line,column=1).font=Font(size=12,color="00FF0000",name="Calibri")
                    sheet.cell(row=line,column=1).alignment=Alignment('center')
                    sheet.cell(row=line,column=1).value=stu_options[m]

            else:
                if(stu_options[m]==correct_options[m]):
                    sheet.cell(row=line-25,column=4).font=Font(size=12,color="00008000",name="Calibri")
                    sheet.cell(row=line-25,column=4).alignment=Alignment('center')
                    sheet.cell(row=line-25,column=4).value=stu_options[m]
                else:
                    sheet.cell(row=line-25,column=4).font=Font(size=12,color="00FF0000",name="Calibri")
                    sheet.cell(row=line-25,column=4).alignment=Alignment('center')
                    sheet.cell(row=line-25,column=4).value=stu_options[m]
            line+=1
            m+=1
        no_of_response+=1

    wb.save(f'output\\{master_roll[no_stu_master]}.xlsx')
    no_stu_master+=1

response_file.insert(5,'score_after_negative',final_score,True)
response_file['Status']=status
response_file.to_excel('output\concise_marksheet.xlsx')

        

        








    
    





